import datetime
import pytz
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (Application, CommandHandler, MessageHandler, filters,
                          ConversationHandler, ContextTypes,
                          CallbackQueryHandler)
import openpyxl
import os
import jdatetime

from apscheduler.schedulers.asyncio import AsyncIOScheduler
import pytz

# Initialize the scheduler
scheduler = AsyncIOScheduler(timezone=pytz.timezone(
    'Asia/Tehran')  # Change 'UTC' to your desired timezone
                             )

# Your other setup code goes here
from datetime import timedelta

# تنظیمات
TOKEN = "7265240527:AAGzXhj6cOkVL8xCnaMlA_c9Tz4gkgA_2Wc"  # توکن جدید از BotFather بگیر
ADMIN_CHAT_ID = -1002393195820
EXCEL_FILE = "payments.xlsx"
RECEIPT_DIR = "receipts"

# مراحل مکالمه
NAME, PHONE, AMOUNT, DATE, FILE, PHONE_HISTORY = range(6)

# منوی اصلی
main_menu = ReplyKeyboardMarkup([[KeyboardButton("📟 ارسال فیش پرداختی")],
                                 [KeyboardButton("📄 تاریخچه پرداخت")]],
                                resize_keyboard=True)

back_button = ReplyKeyboardMarkup([[KeyboardButton("🔙 بازگشت به صفحه اول")]],
                                  resize_keyboard=True)


async def welcome(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    user_first_name = update.effective_user.first_name
    await update.message.reply_text(
        f"👋 خوش آمدید {user_first_name}! یکی از گزینه های زیر را انتخاب کنید:",
        reply_markup=main_menu)
    return ConversationHandler.END


async def start_payment_process(update: Update,
                                context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        " 👤 لطفا نام و نام خانوادگی خود را وارد کنید:",
        reply_markup=back_button)
    return NAME


async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["name"] = update.message.text
    if update.message.text == "🔙 بازگشت به صفحه اول":
        return await welcome(update, context)

    await update.message.reply_text(
        " 📱 لطفا شماره تماس را وارد کنید:\n"
        "در ارسال شماره تماس دقت فرمایید. بررسی تاریخچه پرداختی با شماره تماس شما امکان‌پذیر خواهد بود و همواره یک شماره تماس ثابت و درست را وارد کنید.",
        reply_markup=back_button)
    return PHONE


async def get_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["phone"] = update.message.text
    keyboard = [[InlineKeyboardButton("یک ماه", callback_data="یک ماه")],
                [InlineKeyboardButton("دو ماه", callback_data="دو ماه")],
                [InlineKeyboardButton("سه ماه", callback_data="سه ماه")]]
    if update.message.text == "🔙 بازگشت به صفحه اول":
        return await welcome(update, context)

    await update.message.reply_text(
        " 💰 لطفا مبلغ واریزی را انتخاب کنید:",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return AMOUNT


async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["amount"] = query.data
    await query.edit_message_text(f"💰 مبلغ انتخاب شده: {query.data}")
    await query.message.reply_text(
        " 📅 لطفا تاریخ واریز را وارد کنید (مثلا 1403/03/01):",
        reply_markup=back_button)
    return DATE


async def get_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🔙 بازگشت به صفحه اول":
        return await welcome(update, context)

    date_text = update.message.text.strip()

    # بررسی فرمت و اعتبار تاریخ شمسی
    try:
        # تلاش برای تبدیل به تاریخ شمسی واقعی
        parts = date_text.split("/")
        if len(parts) != 3:
            raise ValueError("فرمت اشتباه")

        year, month, day = map(int, parts)
        jdatetime.date(year, month, day)  # اگر اشتباه باشه این خط خطا می‌ده

    except Exception:
        await update.message.reply_text(
            "❗ فرمت تاریخ صحیح نیست. لطفاً به صورت `yyyy/mm/dd` وارد کنید (مثلاً 1403/03/01):",
            reply_markup=back_button)
        return DATE  # دوباره ازش تاریخ بخواه

    # اگر معتبر بود:
    context.user_data["date"] = date_text
    await update.message.reply_text(" 📟 لطفا تصویر یا فایل فیش را ارسال کنید:",
                                    reply_markup=back_button)
    return FILE


async def get_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "🔙 بازگشت به صفحه اول":
        return await welcome(update, context)

    if update.message.photo:
        file = update.message.photo[-1]
    elif update.message.document:
        file = update.message.document
    elif update.message.text:
        # اگر کاربر پیام متنی فرستاد
        await update.message.reply_text(
            "❗ لطفاً فقط تصویر یا فایل فیش را ارسال کنید.",
            reply_markup=back_button)
        return FILE
    else:
        await update.message.reply_text(
            "❗ لطفاً یک تصویر یا فایل معتبر ارسال کنید.",
            reply_markup=back_button)
        return FILE

    if not os.path.exists(RECEIPT_DIR):
        os.makedirs(RECEIPT_DIR)

    file_path = os.path.join(
        RECEIPT_DIR,
        f"receipt_{update.message.from_user.id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
    )
    file_obj = await context.bot.get_file(file.file_id)
    await file_obj.download_to_drive(file_path)
    context.user_data["file"] = file_path

    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Amount", "Date", "File"])
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        context.user_data["name"], context.user_data["phone"],
        context.user_data["amount"], context.user_data["date"], file_path
    ])
    wb.save(EXCEL_FILE)

    caption = (f"📟 فیش واریزی جدید:\n"
               f"👤 نام: {context.user_data['name']}\n"
               f"📞 شماره: {context.user_data['phone']}\n"
               f"💰 مبلغ: {context.user_data['amount']}\n"
               f"📅 تاریخ: {context.user_data['date']}")

    keyboard = InlineKeyboardMarkup([[
        InlineKeyboardButton(
            "✅ تایید فیش",
            callback_data=f"approve:{context.user_data['phone']}")
    ]])

    with open(file_path, "rb") as f:
        await context.bot.send_photo(chat_id=ADMIN_CHAT_ID,
                                     photo=f,
                                     caption=caption,
                                     reply_markup=keyboard)

    await update.message.reply_text(
        "✅ اطلاعات شما با موفقیت ثبت شد. سپاسگزاریم!", reply_markup=main_menu)
    return ConversationHandler.END


async def approve_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data.split(":")
    if len(data) == 2 and data[0] == "approve":
        phone = data[1]
        await query.edit_message_reply_markup(reply_markup=None)
        await query.message.reply_text(
            f"✅ فیش مربوط به شماره {phone} تایید شد.")


async def ask_for_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("📱 لطفا شماره دانش‌آموزی خود را وارد کنید:"
                                    )
    return PHONE_HISTORY


async def handle_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = update.message.text
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("❗ فایل گزارش پیدا نشد.")
        return ConversationHandler.END

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    results = [
        f"📟 {row[2]} در تاریخ {row[3]}"
        for row in ws.iter_rows(min_row=2, values_only=True) if row[1] == phone
    ]
    msg = f"📄 تاریخچه پرداخت ({phone}):\n\n" + '\n'.join(
        results) if results else "❗ پرداختی با این شماره پیدا نشد."
    await update.message.reply_text(msg, reply_markup=main_menu)
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    return ConversationHandler.END


async def restart_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await cancel(update, context)
    return await start_payment_process(update, context)


async def restart_home(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await cancel(update, context)
    return await welcome(update, context)


async def restart_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await cancel(update, context)
    return await ask_for_phone(update, context)


async def send_excel_to_admin(update: Update,
                              context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.id != ADMIN_CHAT_ID:
        await update.message.reply_text("⛔ شما مجاز به دریافت فایل نیستید.")
        return

    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("❗ فایل پرداختی وجود ندارد.")
        return

    await update.message.reply_document(document=open(EXCEL_FILE, "rb"))


async def send_monthly_text_report(context: ContextTypes.DEFAULT_TYPE):
    today_shamsi = jdatetime.date.today()

    if today_shamsi.day != 1:
        return  # فقط روز اول ماه شمسی اجرا شود

    year = today_shamsi.year
    prev_month = today_shamsi.month - 1 or 12
    prev_year = year if today_shamsi.month != 1 else year - 1

    month_prefix = f"{prev_year}/{str(prev_month).zfill(2)}"  # مثال: 1403/02

    if not os.path.exists(EXCEL_FILE):
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, phone, _, date_str, _ = row
        if date_str.startswith(month_prefix):
            results.append(f"👤 {name} - 📞 {phone}")

    if results:
        message = f"📊 لیست پرداختی‌های ماه {month_prefix}:\n\n" + "\n".join(
            results)
    else:
        message = f"📊 در ماه {month_prefix} هیچ پرداختی ثبت نشده است."

    await context.bot.send_message(chat_id=ADMIN_CHAT_ID, text=message)


async def send_payment_reminders(context: ContextTypes.DEFAULT_TYPE):
    if not os.path.exists(EXCEL_FILE):
        return

    today = jdatetime.date.today()
    target_date = today + timedelta(days=2)  # دو روز بعد

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    notified = set()  # جلوگیری از پیام تکراری

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, phone, _, date_str, _ = row

        try:
            year, month, day = map(int, date_str.split("/"))
            payment_date = jdatetime.date(year, month, day)
        except:
            continue  # اگر تاریخ مشکل داشت، ردش کن

        if payment_date == target_date and phone not in notified:
            notified.add(phone)
            try:
                await context.bot.send_message(
                    chat_id=phone,
                    text=
                    f"⏰ یادآوری:\n{ name } عزیز، فقط دو روز تا موعد پرداخت شما باقی مانده است. لطفاً فیش پرداخت خود را ارسال نمایید.\nسپاس از همکاری شما ❤️"
                )
            except Exception as e:
                print(f"خطا در ارسال پیام به {phone}: {e}")


iran_tz = pytz.timezone("Asia/Tehran")


def main():
    app = Application.builder().token(TOKEN).build()

    # ✅ یادآوری پرداخت روزانه ساعت 8:00 صبح به وقت ایران
    app.job_queue.run_daily(send_payment_reminders,
                            time=datetime.datetime.now(iran_tz).time(),
                            name="daily_reminder")

    # ✅ گزارش ماهانه در روز اول هر ماه، ساعت 00:05 صبح به وقت ایران
    app.job_queue.run_monthly(send_monthly_text_report,
                              when=datetime.datetime.now(iran_tz).time(),
                              day=1,
                              name="monthly_report")

    # بقیه هندلرها و app.run_polling() مثل قبل...

    history_handler = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^📄 تاریخچه پرداخت$"), ask_for_phone)
        ],
        states={
            PHONE_HISTORY:
            [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_history)],
        },
        fallbacks=[CommandHandler("cancel", cancel)])
    app.add_handler(history_handler)
    conv_handler = ConversationHandler(
        entry_points=[
            CommandHandler("start", welcome),
            MessageHandler(filters.Regex("^📟 ارسال فیش پرداختی$"),
                           restart_payment),
        ],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
            PHONE:
            [MessageHandler(filters.TEXT & ~filters.COMMAND, get_phone)],
            AMOUNT: [CallbackQueryHandler(button_handler)],
            DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_date)],
            FILE:
            [MessageHandler(filters.PHOTO | filters.Document.ALL, get_file)],
            PHONE_HISTORY:
            [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_history)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("start", welcome),
            MessageHandler(filters.Regex("^📟 ارسال فیش پرداختی$"),
                           restart_payment),
        ],
        per_message=False)

    app.add_handler(
        MessageHandler(filters.Regex("^📄 تاریخچه پرداخت$"), ask_for_phone))
    app.add_handler(CommandHandler("get_excel", send_excel_to_admin))

    app.add_handler(conv_handler)
    app.add_handler(CallbackQueryHandler(approve_handler, pattern="^approve:"))

    app.add_handler(
        MessageHandler(filters.Regex("^🔙 بازگشت به صفحه اول$"), welcome))

    app.run_polling()


if __name__ == "__main__":
    main()
